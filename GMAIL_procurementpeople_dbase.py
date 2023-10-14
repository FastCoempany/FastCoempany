import openpyxl

# Load the workbook and select the active sheet
input_path = "/Users/antaeus.coe/Desktop/dbase_slack_alerts/Parent scraped file for dbase alerts.xlsx"
wb = openpyxl.load_workbook(input_path)
sheet = wb.active

# Create a set to keep track of seen entries to remove duplicates
seen = set()

# Create a new workbook for the cleaned data
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# Define the headers for the new sheet
headers = ["First Name", "Last Name", "Title", "Company"]
new_sheet.append(headers)

# Iterate through the rows in the original sheet
for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming row 1 has headers
    if row[0]:  # Check if the name column is not None
        if '_' in row[0]:  # Check if the name has an underscore
            first_name, last_name = row[0].split('_')
        else:
            name_parts = row[0].split()
            first_name = name_parts[0]
            last_name = ' '.join(name_parts[1:])
        title = row[1]
        company = row[2]

        # Check if this entry is a duplicate
        if (first_name, last_name, title, company) not in seen:
            seen.add((first_name, last_name, title, company))
            new_sheet.append([first_name, last_name, title, company])

# Save the new workbook
output_path = "/Users/antaeus.coe/Desktop/dbase_slack_alerts/GMAIL_procurementpeople_dbase.xlsx"
new_wb.save(output_path)

print(f"Data saved to {output_path}")
