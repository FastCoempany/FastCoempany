import openpyxl
import openai

# Initialize OpenAI key
openai.api_key = "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"  # Replace with your actual key

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook("/Users/antaeus.coe/Desktop/dbase_slack_alerts/Unique_URL_3.xlsx")
sheet = wb.active

def generate_email_body(url, description, opinion):
    # Construct the input prompt for ChatGPT
    prompt = f"Write a down-to-earth, attention-grabbing, professional email body about the topic '{description}' with the opinion '{opinion}', with a 60-word limit."
    
    # Request a completion from ChatGPT
    response = openai.Completion.create(engine="text-davinci-002", prompt=prompt, max_tokens=100)
    
    # Get the content from the response
    email_body = response.choices[0].text.strip()
    
    # Create hyperlink for URL
    link_text = description.split()[0]
    hyperlink = f'You visited <a href="{url}">{link_text}</a>.'
    
    # Construct the final email body with appended message and hyperlink
    final_email_body = f"{hyperlink} {email_body} My initial conversations are kept to 15 minutes. If you'd like - we can even [start with a virtual 30 seconds here right now](https://fiverrent.wistia.com/medias/tpqf6eb18f) - and you can decide for yourself at the end if you'd like to have the 15min chat live."
    
    # Embed a clandestine signal within the email body for automation tool detection
    signal = "\n\n#sig#"
    email_body_with_signal = f"{final_email_body} {signal}"
    return email_body_with_signal

for i in range(2, sheet.max_row + 1):
    url = sheet.cell(row=i, column=1).value
    description = sheet.cell(row=i, column=2).value
    opinion = sheet.cell(row=i, column=3).value
    
    email_body = generate_email_body(url, description, opinion)
    # Write the generated email bodies to another column 
    sheet.cell(row=i, column=4).value = email_body

# Save the workbook
wb.save("/Users/antaeus.coe/Desktop/dbase_slack_alerts/Unique_URL_3_updated.xlsx")
