import openpyxl
import openai
import os.path
import pickle
import random
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from google.auth.transport.requests import Request

# 1. Initialize OpenAI key
openai.api_key = "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"

# 2. Load the data workbook
data_wb = openpyxl.load_workbook("/Users/antaeus.coe/Desktop/dbase_slack_alerts/organized again alerts.xlsx")
data_sheet = data_wb.active

# 3. Load URL content data into a dictionary for easy access and another for company-URL mapping
url_content = {}
company_urls = {}

DEFAULT_URL = "https://enterprise.fiverr.com/webinars/how-to-successfully-manage-a-thriving-freelance-workforce/"
DEFAULT_DESCRIPTION = "A collection of e-books provided by Fiverr Enterprise on various topics related to freelancing and business."
DEFAULT_OPINION = "A treasure trove of in-depth insights for businesses to dive deep into freelance management best practices and strategies."

# Loop over all rows in the file, starting from the second row (assuming the first row contains headers)
for i in range(2, data_sheet.max_row + 1):
    url = data_sheet.cell(row=i, column=5).value
    description = data_sheet.cell(row=i, column=6).value
    opinion = data_sheet.cell(row=i, column=7).value

    company = str(data_sheet.cell(row=i, column=4).value).strip()
    unique_url = data_sheet.cell(row=i, column=8).value
    url_content[url] = (description, opinion)

    if company in company_urls:
        if unique_url not in company_urls[company]:
            company_urls[company].append(unique_url)
    else:
        company_urls[company] = [unique_url]

    first_name = data_sheet.cell(row=i, column=1).value
    last_name = data_sheet.cell(row=i, column=2).value
    company = str(data_sheet.cell(row=i, column=4).value).strip()

# Debug: Print company and its associated URLs
print(f"Company: {company}")
print(f"URLs associated with the company: {company_urls.get(company, 'No URLs found')}")

# 4. Email generation function
def generate_email_body(first_name, last_name, company, url_content, company_urls):
    
    # Check for initials in first name
    greeting = last_name if len(first_name) <= 2 else first_name

    # If the company is not in the company_urls dictionary or has no URLs associated, use the default URL, Description, and Opinion
    if company not in company_urls or not company_urls[company]:
        chosen_url = DEFAULT_URL
        description = DEFAULT_DESCRIPTION
        opinion = DEFAULT_OPINION
    else:
        # Choose a random URL for companies with multiple URLs
        visited_urls = company_urls[company]
        chosen_url = random.choice(visited_urls) if len(visited_urls) > 1 else visited_urls[0]
        description, opinion = url_content.get(chosen_url, (DEFAULT_DESCRIPTION, DEFAULT_OPINION))

    # Construct the input prompt for GPT-4
    prompt_text = (f"Based on the unique URL {chosen_url} they visited, Craft a concise message highlighting the benefits of Fiverr Enterprise. "
                   f"Emphasize its freelancer management capabilities, especially in finance, payment, admin, compliance, legal, and global sourcing. "
                   f"Refer to the URL's context: '{description}' - '{opinion}'. Make it relevant for procurement executives. Ensure completeness and a concluding period.")
    
    response = openai.Completion.create(engine="text-davinci-002", prompt=prompt_text, max_tokens=80)
    email_body_portion = response.choices[0].text.strip()

    # Starting the first paragraph with 'We saw you visited ...'
    email_body = (f"Hi {greeting},\n\n"
                  f"We saw you visited [{description}]({chosen_url}). {email_body_portion}\n\n"
                  f"My initial conversations are kept to 15 minutes. If you're interested, we can "
                  f"[start with a virtual 30 seconds here right now](https://fiverrent.wistia.com/medias/tpqf6eb18f) "
                  f"and then decide if a 15-min chat would be beneficial.\n\n\n\n")
    
    return email_body


# 5. Google Docs API setup and authentication
SCOPES = ['https://www.googleapis.com/auth/documents']
creds = None

if os.path.exists('token_k.pkl'):
    with open('token_k.pkl', 'rb') as token:
        creds = pickle.load(token)

if not creds or not creds.valid:
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    else:
        flow = InstalledAppFlow.from_client_secrets_file(
            '/Users/antaeus.coe/Desktop/dbase_slack_alerts/client_secret_362571091781-18hhrut49ku1t1gek1ml6922ik0fge87.apps.googleusercontent.com.json', SCOPES)
        creds = flow.run_local_server(port=0)
    with open('token_k.pkl', 'wb') as token:
        pickle.dump(creds, token)

doc_service = build('docs', 'v1', credentials=creds)
document_response = doc_service.documents().create(body={'title': 'Generated Emails-I'}).execute()
DOCUMENT_ID = document_response.get('documentId')
print(f"Document ID: {DOCUMENT_ID}")

if not DOCUMENT_ID:
    print("Error: Unable to create a new Google document or retrieve its documentId.")
    exit(1)

for i in range(2, data_sheet.max_row + 1):
    first_name = data_sheet.cell(row=i, column=1).value
    last_name = data_sheet.cell(row=i, column=2).value
    company = data_sheet.cell(row=i, column=4).value

    email_body = generate_email_body(first_name, last_name, company, url_content, company_urls)

    requests = [{
        'insertText': {
            'location': {
                'index': 1,
            },
            'text': email_body + ".#howboutahighfiverr#."
        }
    }]
    
    doc_service.documents().batchUpdate(documentId=DOCUMENT_ID, body={'requests': requests}).execute()





