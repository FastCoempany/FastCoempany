import openpyxl

# Load the workbook with unorganized data
src_wb = openpyxl.load_workbook("/Users/antaeus.coe/Desktop/dbase_slack_alerts/Dbase unorganized Alerts.xlsx")
src_sheet = src_wb.active

# Create a new workbook for organized data
dst_wb = openpyxl.Workbook()
dst_sheet = dst_wb.active

# Set headers for the new sheet
dst_sheet["A1"] = "Company"
dst_sheet["B1"] = "Web pages read"

dst_row = 2  # Start adding organized data from the second row in the new sheet

# Iterate through the rows in the source sheet
row = 1
while row <= src_sheet.max_row:
    cell_value = str(src_sheet.cell(row=row, column=1).value)  # Convert cell value to string

    if "Web pages read:" in cell_value:
        # Extract company name
        company = cell_value.split("Web pages read:")[0].strip()
        
        # Increment row to start reading URLs
        row += 1
        while row <= src_sheet.max_row and src_sheet.cell(row=row, column=1).value != "Preferences":
            # Write company name and URL to the new sheet
            dst_sheet.cell(row=dst_row, column=1, value=company)
            dst_sheet.cell(row=dst_row, column=2, value=src_sheet.cell(row=row, column=1).value)
            dst_row += 1
            
            # Move to the next row to continue checking for more URLs
            row += 1

    # Move to the next row
    row += 1

# Save the organized data to a new file
dst_wb.save("/Users/antaeus.coe/Desktop/dbase_slack_alerts/organized again alerts.xlsx")

