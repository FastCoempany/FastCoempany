import openpyxl

# Load workbook
workbook = openpyxl.load_workbook('Unique_URL_3_updated.xlsx')

# Define the source and destination sheets
source_sheet = workbook['Sheet1']
dest_sheet = workbook['Organized Data']

# Determine where to start writing in the destination sheet
dest_row = 71

# Loop through the source sheet in sets of 4 rows (URL, description, opinion, blank)
for i in range(1, source_sheet.max_row + 1, 4):
    url = source_sheet.cell(row=i, column=1).value
    description = source_sheet.cell(row=i+1, column=1).value
    opinion = source_sheet.cell(row=i+2, column=1).value

    # Write the extracted data to the destination sheet
    dest_sheet.cell(row=dest_row, column=1, value=url)
    dest_sheet.cell(row=dest_row, column=2, value=description)
    dest_sheet.cell(row=dest_row, column=3, value=opinion)

    # Move to the next row in the destination sheet for the next set of data
    dest_row += 1

# Save the updated workbook
workbook.save('Unique_URL_3_updated.xlsx')



