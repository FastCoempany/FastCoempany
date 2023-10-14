import openpyxl

file_path = "/Users/antaeus.coe/Desktop/dbase_slack_alerts/backup.modified_messagesfinal710pm.xlsx"

# Load the workbook and select the sheet
book = openpyxl.load_workbook(file_path)
ws = book['blueblocker']

# Collect rows that are not blank
rows_to_keep = [row for row in ws.iter_rows() if any(cell.value for cell in row)]

# Clear the sheet
for row in ws.iter_rows():
    for cell in row:
        cell.value = None

# Write back the non-blank rows to the sheet
for r_idx, row in enumerate(rows_to_keep, 1):
    for c_idx, cell in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=cell.value)

# Save the changes
book.save(file_path)

