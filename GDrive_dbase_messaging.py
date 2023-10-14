import openpyxl
import openai
import os.path
import pickle
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from google.auth.transport.requests import Request

# Initialize OpenAI key
openai.api_key = "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"  # Make sure to insert your actual key here

# Load the workbook and select the active sheet
wb = openpyxl.load_workbook("/Users/antaeus.coe/Desktop/dbase_slack_alerts/Unique_URL_3.xlsx")
sheet = wb.active

def generate_email_body(url, description, opinion):
    # Construct the input prompt for ChatGPT
    prompt = (f"Write an informative and prescriptive down-to-earth, attention-grabbing, professional email body "
              f"starting with 'We saw you visited [{{hyperlink text based on the URL}}]({url})'. "
              f"Discuss the topic '{description}' and the opinion '{opinion}'. The response should adhere to a 60-word limit.")
    
    # Request a completion from ChatGPT
    response = openai.Completion.create(engine="text-davinci-002", prompt=prompt, max_tokens=100)
    
    # Get the content from the response
    email_body = response.choices[0].text.strip()
    
    # Append the default message and tag
    email_body += (f"\n\nMy initial conversations are kept to 15 minutes. If you'd like - we can "
                   f"even [start with a virtual 30 seconds here right now]"
                   f"(https://fiverrent.wistia.com/medias/tpqf6eb18f) - and you can decide for "
                   f"yourself at the end if you'd like to have the 15min chat live.\n\n\n"
                   f".#howboutahighfiverr#.\n\n\n")
    
    return email_body

# Google Docs API setup and authentication
SCOPES = ['https://www.googleapis.com/auth/documents']

# Check if token_k.pkl exists
creds = None
if os.path.exists('token_k.pkl'):
    with open('token_k.pkl', 'rb') as token:
        creds = pickle.load(token)

# If there are no valid credentials available, ask the user to log in.
if not creds or not creds.valid:
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    else:
        flow = InstalledAppFlow.from_client_secrets_file('/Users/antaeus.coe/Desktop/dbase_slack_alerts/client_secret_362571091781-18hhrut49ku1t1gek1ml6922ik0fge87.apps.googleusercontent.com.json', SCOPES)
        creds = flow.run_local_server(port=0)
        
    # Save the credentials for the next run
    with open('token_k.pkl', 'wb') as token:
        pickle.dump(creds, token)

doc_service = build('docs', 'v1', credentials=creds)

# Create a new Google Document and retrieve its ID
document_response = doc_service.documents().create(body={'title': 'Generated Emails6'}).execute()
DOCUMENT_ID = document_response.get('documentId')

if not DOCUMENT_ID:
    print("Error: Unable to create a new Google document or retrieve its documentId.")
    exit(1)

for i in range(2, sheet.max_row + 1):
    url = sheet.cell(row=i, column=1).value
    description = sheet.cell(row=i, column=2).value
    opinion = sheet.cell(row=i, column=3).value
    
    email_body = generate_email_body(url, description, opinion)
    
    # Add the generated email body to the Google Doc
    requests = [
        {
            'insertText': {
                'location': {
                    'index': 1,
                },
                'text': email_body + '\n\n\n'
            }
        }
    ]

    doc_service.documents().batchUpdate(documentId=DOCUMENT_ID, body={'requests': requests}).execute()



