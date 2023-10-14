import openpyxl

# Open the workbook and select the active sheet
wb = openpyxl.load_workbook("/Users/antaeus.coe/Desktop/dbase_slack_alerts/Unique_URL_3.xlsx")
sheet = wb.active

# Create a new sheet for organized data
new_sheet = wb.create_sheet(title="Organized Data")

# Set headers in the new sheet
new_sheet["A1"] = "URL"
new_sheet["B1"] = "Description"
new_sheet["C1"] = "Opinion"

# Iterate through cells in the original sheet and populate the new sheet
row_num = 2
for i in range(1, sheet.max_row + 1, 5):
    # Get the URL, Description, and Opinion cells' values
    url = sheet.cell(row=i, column=1).value
    description = sheet.cell(row=i + 2, column=1).value.replace("Description:", "").strip()
    opinion = sheet.cell(row=i + 3, column=1).value.replace("Opinion:", "").strip()

    # Write values to the new sheet
    new_sheet.cell(row=row_num, column=1).value = url
    new_sheet.cell(row=row_num, column=2).value = description
    new_sheet.cell(row=row_num, column=3).value = opinion
    row_num += 1

# Save the workbook
wb.save("/Users/antaeus.coe/Desktop/dbase_slack_alerts/Unique_URL_3.xlsx")



